"""
Attendance Template Writer - Generates Excel attendance sheets.
Refactored to use OOP principles with base classes for better reusability.
"""

import calendar
from datetime import date

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domain.employee import Employee
from infrastructure.excel_template_writer import ExcelTemplateWriter
from utils.calendar_utils import get_arabic_month, get_weeks, to_arabic_ordinal


class AttendanceTemplateWriter(ExcelTemplateWriter):
    """
    Writes attendance templates to Excel files.
    Inherits common functionality from ExcelTemplateWriter.
    """

    def __init__(self, output_directory: str = "./output"):
        """
        Initialize the attendance template writer.

        Args:
            output_directory: Directory where Excel files will be saved
        """
        super().__init__(output_directory)

    def _prepare_employees_list(self, template):
        """
        Combine employees and stylists into a single list.

        Args:
            template: AttendanceTemplate object

        Returns:
            Combined list of employees and stylists as Employee objects
        """
        stylists_as_employees = [
            Employee(
                name=s.name,
                daily_rate=s.daily_rate,
                working_hours=s.working_hours,
            )
            for s in template.stylists
        ]
        return template.employees + stylists_as_employees

    def _write_header(self, ws: Worksheet, template, num_employees: int):
        """
        Write the header section of the attendance sheet.

        Args:
            ws: Worksheet to write to
            template: AttendanceTemplate object
            num_employees: Total number of employees (including stylists)
        """
        # Merge cells for title
        ws.merge_cells("A2:B3")
        ws.merge_cells(f"C2:{get_column_letter(num_employees + 2)}2")

        # Title
        ws["C2"] = (
            f"كشف حضور و إنصراف الموظفين عن شهر {get_arabic_month(template.month)} {template.year}"
        )
        ws["C2"].fill = self.style_manager.primary_fill
        ws["C2"].font = self.style_manager.arial_bold_font

        # Date label
        ws["A2"] = "التاريخ"
        ws["A2"].fill = self.style_manager.primary_fill
        ws["A2"].font = self.style_manager.arial_bold_font

    def _write_employee_names(self, ws: Worksheet, employees, row: int):
        """
        Write employee names in the header row.

        Args:
            ws: Worksheet to write to
            employees: List of employees
            row: Row number to write to
        """
        col = 3
        for emp in employees:
            ws.cell(row=row, column=col).value = emp.name
            ws.cell(row=row, column=col).font = self.style_manager.dubai_bold_font
            col += 1

    def _write_day_rows(
        self,
        ws: Worksheet,
        week,
        template,
        employees,
        starting_row: int,
    ):
        """
        Write rows for each day in a week.

        Args:
            ws: Worksheet to write to
            week: List of day numbers in the week
            template: AttendanceTemplate object
            employees: List of employees
            starting_row: Starting row number

        Returns:
            Tuple of (next_row, list_of_day_total_rows)
        """
        row = starting_row
        col = 1
        cells_for_each_day = ["وقت الحضور", "وقت الإنصراف", "ملاحظات", "إحتساب اليوم"]
        each_day_total_row = []

        for _, day in enumerate(week):
            # Merge cells for date
            ws.merge_cells(
                start_row=row,
                start_column=col,
                end_row=row + len(cells_for_each_day) - 1,
                end_column=col,
            )

            # Write date
            d = date(year=template.year, month=template.month, day=day)
            ws[self.cell(col, row)].value = d
            ws[self.cell(col, row)].font = self.style_manager.dubai_bold_font
            ws[self.cell(col, row)].number_format = "dddd, mmmm d, yyyy"

            # Write day detail rows
            for i, cell_label in enumerate(cells_for_each_day):
                ws[self.cell(col + 1, row + i)].value = cell_label
                ws[self.cell(col + 1, row + i)].font = (
                    self.style_manager.dubai_bold_font
                )

                # Set time format for presence and departure cells
                if cell_label in ("وقت الحضور", "وقت الإنصراف"):
                    for m in range(len(employees)):
                        col_p = col + 2 + m
                        ws[self.cell(row=row + i, col=col_p)].number_format = (
                            "h:mm AM/PM"
                        )

                # Add formula for day calculation
                if cell_label == "إحتساب اليوم":
                    for m, emp in enumerate(employees):
                        col_p = col + 2 + m
                        ws[self.cell(row=row + i, col=col_p)].value = (
                            f"=({self.cell(col=col_p, row=row+1)}-{self.cell(col=col_p, row=row)}+IF({self.cell(col=col_p, row=row+1)}<{self.cell(col=col_p, row=row)},1,0))*24/{emp.working_hours}"
                        )
                        ws[self.cell(row=row + i, col=col_p)].font = (
                            self.style_manager.dubai_bold_font
                        )

            each_day_total_row.append(row + 3)
            row += len(cells_for_each_day)

        return row, each_day_total_row

    def _write_week_total(
        self,
        ws: Worksheet,
        week_number: int,
        row: int,
        col: int,
        employees,
        each_day_total_row,
    ):
        """
        Write the week total row.

        Args:
            ws: Worksheet to write to
            week_number: Week number (1-indexed)
            row: Row number to write to
            col: Starting column
            employees: List of employees
            each_day_total_row: List of row numbers containing day totals

        Returns:
            Row number of the week total
        """
        # Week total formulas for each employee
        for i, emp in enumerate(employees):
            formula = "="
            col_p = col + 2 + i
            for m, total_row in enumerate(each_day_total_row):
                formula += self.cell(col=col_p, row=total_row)
                if m != len(each_day_total_row) - 1:
                    formula += "+"
            ws[self.cell(col=col_p, row=row)].value = formula
            ws[self.cell(col=col_p, row=row)].fill = self.style_manager.primary_fill
            ws[self.cell(col=col_p, row=row)].font = self.style_manager.arial_bold_font

        # Week labels
        ws[self.cell(col, row)].value = (
            f"نهاية الأسبوع {to_arabic_ordinal(week_number)}"
        )
        ws[self.cell(col, row)].fill = self.style_manager.primary_fill
        ws[self.cell(col, row)].font = self.style_manager.arial_bold_font

        ws[self.cell(col + 1, row)].value = "إجمالى عدد الايام"
        ws[self.cell(col + 1, row)].fill = self.style_manager.primary_fill
        ws[self.cell(col + 1, row)].font = self.style_manager.arial_bold_font

        return row

    def _write_month_total(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        employees,
        each_week_total_row,
    ):
        """
        Write the month total row.

        Args:
            ws: Worksheet to write to
            row: Row number to write to
            col: Starting column
            employees: List of employees
            each_week_total_row: List of row numbers containing week totals
        """
        # Month total formulas for each employee
        for i, emp in enumerate(employees):
            formula = "="
            col_p = col + 2 + i
            for m, total_row in enumerate(each_week_total_row):
                formula += self.cell(col=col_p, row=total_row)
                if m != len(each_week_total_row) - 1:
                    formula += "+"
            ws[self.cell(col=col_p, row=row)].value = formula
            ws[self.cell(col=col_p, row=row)].fill = self.style_manager.primary_fill
            ws[self.cell(col=col_p, row=row)].font = self.style_manager.arial_bold_font

        # Month labels
        ws[self.cell(col=1, row=row)].fill = self.style_manager.primary_fill
        ws[self.cell(col=1, row=row)].font = self.style_manager.arial_bold_font
        ws[self.cell(col=2, row=row)].fill = self.style_manager.primary_fill
        ws[self.cell(col=2, row=row)].font = self.style_manager.arial_bold_font

        ws[self.cell(col=1, row=row)].value = "نهاية الشهر"
        ws[self.cell(col=2, row=row)].value = "إجمالي حضور الشهر"

    def write(self, template, use_month_folder: bool = False):
        """
        Write the attendance template to an Excel file.

        Args:
            template: AttendanceTemplate object containing all data
            use_month_folder: If True, saves to a month folder instead of directly to output
        """
        # Prepare data
        employees = self._prepare_employees_list(template)
        template.employees = employees

        # Create workbook and worksheet
        wb = self.create_workbook()
        ws: Worksheet = wb.active
        self.worksheet_helper.configure_worksheet(ws, rtl=True, show_grid=False)

        # Apply default font to all cells
        self.worksheet_helper.apply_font_to_all(
            ws, self.style_manager.arial_regular_font
        )

        ws.title = "Attendance"

        # Write header
        self._write_header(ws, template, len(employees))

        # Write employee names
        self._write_employee_names(ws, employees, row=3)

        # Write day rows for each week
        row = 4
        col = 1
        month_weeks = get_weeks(year=template.year, month=template.month)
        each_week_total_row = []

        for j, week in enumerate(month_weeks):
            starting_week_row = row
            row, each_day_total_row = self._write_day_rows(
                ws, week, template, employees, row
            )

            # Write week total
            week_total_row = self._write_week_total(
                ws, j + 1, row, col, employees, each_day_total_row
            )
            each_week_total_row.append(week_total_row)
            row += 1

        # Write month total
        self._write_month_total(ws, row, col, employees, each_week_total_row)

        # Auto-fit and styling
        self.worksheet_helper.auto_fit_sheet(ws)

        # Calculate dimensions for borders
        days_in_month = calendar.monthrange(template.year, template.month)[1]
        max_used_col = 2 + len(employees)
        max_used_row = 4 * days_in_month + len(each_week_total_row) + 1 + 3

        # Add borders
        self.worksheet_helper.add_borders_to_range(
            ws, max_col_idx=max_used_col, max_row_idx=max_used_row
        )

        # # Apply final font styling
        self.apply_fonts_by_fill(ws)

        # Freeze panes
        ws.freeze_panes = "A4"

        # Save workbook
        self.save_workbook(
            wb,
            template.year,
            template.month,
            use_month_folder=use_month_folder,
        )

    def apply_fonts_by_fill(self, ws):
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill == self.style_manager.primary_fill:
                    cell.font = self.style_manager.arial_bold_font
                else:
                    cell.font = self.style_manager.dubai_bold_font
