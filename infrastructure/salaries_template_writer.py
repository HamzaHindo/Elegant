"""
Salaries Template Writer - Generates Excel salary sheets.
Uses Excel formulas to link to attendance data for working hours.
"""

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domain.employee import Employee
from domain.salary_template import SalaryTemplate
from infrastructure.excel_template_writer import ExcelTemplateWriter
from utils.calendar_utils import get_arabic_month, get_weeks, to_arabic_ordinal


class SalariesTemplateWriter(ExcelTemplateWriter):
    """
    Writes salary templates to Excel files.
    Uses Excel formulas to reference attendance data.
    """

    def __init__(self, output_directory: str = "./output"):
        """
        Initialize the salaries template writer.

        Args:
            output_directory: Directory where Excel files will be saved
        """
        super().__init__(output_directory)

    def _prepare_employees_list(self, template: SalaryTemplate):
        """
        Combine employees and stylists into a single list.

        Args:
            template: SalaryTemplate object

        Returns:
            Combined list of employees and stylists as Employee objects
        """
        stylists_as_employees = [
            Employee(
                name=s.name,
                daily_rate=s.daily_rate,
                working_hours=s.working_hours,
            )
            for s in template.stylists
        ]
        return template.employees + stylists_as_employees

    def _get_attendance_formula(
        self, employee_name: str, week_number: int, attendance_filename: str
    ) -> str:
        """
        Generate Excel formula to get working hours from attendance file.

        Args:
            employee_name: Name of the employee
            week_number: Week number (1-indexed)
            attendance_filename: Name of the attendance file

        Returns:
            Excel formula string
        """
        # Formula to lookup employee column and get week total
        # Uses MATCH to find employee column, then INDEX to get the week total value
        # The formula looks for "نهاية الأسبوع {ordinal}" in column A
        week_ordinal = to_arabic_ordinal(week_number)

        # Formula structure:
        # =IFERROR(INDEX('[Attendance_12_2026.xlsx]Attendance'!$C$3:$Z$1000,
        #          MATCH("نهاية الأسبوع الأول",'[Attendance_12_2026.xlsx]Attendance'!$A:$A,0),
        #          MATCH("أحمد محمد",'[Attendance_12_2026.xlsx]Attendance'!$3:$3,0)-2),0)

        formula = (
            f"=IFERROR("
            f"INDEX('[{attendance_filename}]Attendance'!$A$1:$Z$1000,"
            f"MATCH(\"نهاية الأسبوع {week_ordinal}\",'[{attendance_filename}]Attendance'!$A:$A,0),"
            f"MATCH(\"{employee_name}\",'[{attendance_filename}]Attendance'!$3:$3,0)"
            f"),0)"
        )

        return formula

    def _get_borrows_formula(
        self, employee_name: str, week_number: int, borrows_filename: str
    ) -> str:
        """
        Generate Excel formula to get working hours from attendance file.

        Args:
            employee_name: Name of the employee
            week_number: Week number (1-indexed)
            attendance_filename: Name of the attendance file

        Returns:
            Excel formula string
        """
        # Formula to lookup employee column and get week total
        # Uses MATCH to find employee column, then INDEX to get the week total value
        # The formula looks for "نهاية الأسبوع {ordinal}" in column A
        week_ordinal = to_arabic_ordinal(week_number)

        # Formula structure:
        # =IFERROR(INDEX('[Attendance_12_2026.xlsx]Attendance'!$C$3:$Z$1000,
        #          MATCH("نهاية الأسبوع الأول",'[Attendance_12_2026.xlsx]Attendance'!$A:$A,0),
        #          MATCH("أحمد محمد",'[Attendance_12_2026.xlsx]Attendance'!$3:$3,0)-2),0)

        sheet_name = "كشف السلف"

        formula = (
            f"=IFERROR("
            f"INDEX('[{borrows_filename}]{sheet_name}'!$A$1:$Z$1000,"
            f"MATCH(\"نهاية الأسبوع {week_ordinal}\",'[{borrows_filename}]{sheet_name}'!$A:$A,0),"
            f"MATCH(\"{employee_name}\",'[{borrows_filename}]{sheet_name}'!$3:$3,0)"
            f"),0)"
        )

        return formula

    def _write_week_header(
        self,
        ws: Worksheet,
        row: int,
        week_number: int,
        month: int,
        year: int,
    ):
        """
        Write the header for a week section.

        Args:
            ws: Worksheet to write to
            row: Starting row number
            week_number: Week number (1-indexed)
            month: Month number
            year: Year
        """
        # Merge cells from B to J for the header
        ws.merge_cells(f"B{row}:J{row}")

        # Write header text
        header_text = f"رواتب الموظفين عن شهر {get_arabic_month(month)} - الأسبوع {to_arabic_ordinal(week_number)}"
        cell = ws[f"B{row}"]
        cell.value = header_text
        cell.fill = self.style_manager.primary_fill
        cell.font = self.style_manager.header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def _write_column_headers(self, ws: Worksheet, row: int):
        """
        Write column headers for the salary table.

        Args:
            ws: Worksheet to write to
            row: Row number for headers
        """
        headers = [
            "اسم الموظف",  # B - Employee name
            "يومية الموظف",  # C - Daily rate
            "عدد أيام العمل خلال الأسبوع",  # D - Working days
            "الخصومات",  # E - Deductions
            "عدد أيام العمل بعد الخصومات",  # F - Working days after deductions
            "الراتب الأسبوعي",  # G - Weekly salary (without deductions)
            "الراتب الأسبوعي بعد الخصم",  # H - Weekly salary (with deductions)
            "سلف الموظف خلال الأسبوع",  # I - Employee advances
            "نهائي الراتب المدفوع بعد خصم السلف",  # J - Final salary
        ]

        for i, header in enumerate(headers, start=2):  # Start from column B (2)
            cell = ws.cell(row=row, column=i)
            cell.value = header
            cell.fill = self.style_manager.primary_fill
            cell.font = self.style_manager.header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    def _write_employee_row(
        self,
        ws: Worksheet,
        row: int,
        employee: Employee,
        week_number: int,
        attendance_filename: str,
        borrows_filename: str,
    ):
        """
        Write a single employee's salary row.

        Args:
            ws: Worksheet to write to
            row: Row number
            employee: Employee object
            week_number: Week number (1-indexed)
            attendance_filename: Name of the attendance file
        """
        # B - Employee name
        ws.cell(row=row, column=2).value = employee.name
        ws.cell(row=row, column=2).font = self.style_manager.bold_font
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="right", vertical="center"
        )

        # C - Daily rate
        ws.cell(row=row, column=3).value = employee.daily_rate
        ws.cell(row=row, column=3).font = self.style_manager.regular_font
        ws.cell(row=row, column=3).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=3).number_format = "#,##0.00"

        # D - Working days (Excel formula to get from attendance)
        formula = self._get_attendance_formula(
            employee.name, week_number, attendance_filename
        )
        ws.cell(row=row, column=4).value = formula
        ws.cell(row=row, column=4).font = self.style_manager.regular_font
        ws.cell(row=row, column=4).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=4).number_format = "#,##0.00"

        # E - Deductions (default 0)
        ws.cell(row=row, column=5).value = 0
        ws.cell(row=row, column=5).font = self.style_manager.regular_font
        ws.cell(row=row, column=5).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=5).number_format = "#,##0.00"

        # F - Working days after deductions (D - E)
        ws.cell(row=row, column=6).value = f"=D{row}-E{row}"
        ws.cell(row=row, column=6).font = self.style_manager.regular_font
        ws.cell(row=row, column=6).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=6).number_format = "#,##0.00"

        # G - Weekly salary without deductions (C * D)
        ws.cell(row=row, column=7).value = f"=C{row}*D{row}"
        ws.cell(row=row, column=7).font = self.style_manager.regular_font
        ws.cell(row=row, column=7).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=7).number_format = "#,##0.00"

        # H - Weekly salary with deductions (C * F)
        ws.cell(row=row, column=8).value = f"=C{row}*F{row}"
        ws.cell(row=row, column=8).font = self.style_manager.regular_font
        ws.cell(row=row, column=8).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=8).number_format = "#,##0.00"

        formula = self._get_borrows_formula(
            employee.name, week_number, borrows_filename
        )

        # I - Employee advances (default 0, can be edited)
        ws.cell(row=row, column=9).value = formula
        ws.cell(row=row, column=9).font = self.style_manager.regular_font
        ws.cell(row=row, column=9).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=9).number_format = "#,##0.00"

        # J - Final salary (H - I)
        ws.cell(row=row, column=10).value = f"=H{row}-I{row}"
        ws.cell(row=row, column=10).font = self.style_manager.bold_font
        ws.cell(row=row, column=10).fill = self.style_manager.secondary_fill
        ws.cell(row=row, column=10).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=10).number_format = "#,##0.00"

    def _write_week_total(
        self, ws: Worksheet, row: int, first_data_row: int, last_data_row: int
    ):
        """
        Write the total row for a week.

        Args:
            ws: Worksheet to write to
            row: Row number for the total
            first_data_row: First row with employee data
            last_data_row: Last row with employee data
        """
        # Merge B to I for the label
        ws.merge_cells(f"B{row}:I{row}")
        ws[f"B{row}"].value = "إجمالي الرواتب للأسبوع"
        ws[f"B{row}"].fill = self.style_manager.primary_fill
        ws[f"B{row}"].font = self.style_manager.header_font
        ws[f"B{row}"].alignment = self.style_manager.center_alignment

        # J - Sum of all final salaries
        ws.cell(row=row, column=10).value = f"=SUM(J{first_data_row}:J{last_data_row})"
        ws.cell(row=row, column=10).fill = self.style_manager.primary_fill
        ws.cell(row=row, column=10).font = self.style_manager.header_font
        ws.cell(row=row, column=10).alignment = self.style_manager.center_alignment
        ws.cell(row=row, column=10).number_format = "#,##0.00"

    def _write_week_section(
        self,
        ws: Worksheet,
        starting_row: int,
        week_number: int,
        employees: list[Employee],
        attendance_filename: str,
        borrows_filename: str,
        template: SalaryTemplate,
    ) -> int:
        """
        Write a complete week section.

        Args:
            ws: Worksheet to write to
            starting_row: Starting row for this section
            week_number: Week number (1-indexed)
            employees: List of all employees
            attendance_filename: Name of the attendance file
            template: SalaryTemplate object

        Returns:
            Next available row number
        """
        row = starting_row

        # Write week header
        self._write_week_header(ws, row, week_number, template.month, template.year)
        row += 1

        # Write column headers
        self._write_column_headers(ws, row)
        row += 1

        # Track first and last data rows for total calculation
        first_data_row = row

        # Write employee rows
        for employee in employees:
            self._write_employee_row(
                ws, row, employee, week_number, attendance_filename, borrows_filename
            )
            row += 1

        last_data_row = row - 1

        # Write week total
        self._write_week_total(ws, row, first_data_row, last_data_row)
        row += 1

        # Add borders to the entire section
        self.worksheet_helper.outline_range(
            ws,
            start_row=starting_row,
            start_col=2,  # Column B
            end_row=row - 1,
            end_col=10,  # Column J
        )

        return row

    def _write_grand_total(
        self,
        ws: Worksheet,
        row: int,
        week_total_rows: list[int],
    ):
        """
        Write the grand total row for all weeks.

        Args:
            ws: Worksheet to write to
            row: Row number for the grand total
            week_total_rows: List of row numbers containing week totals
        """
        # Merge B to I for the label
        ws.merge_cells(f"B{row}:I{row}")
        ws[f"B{row}"].value = "إجمالي الرواتب للشهر بالكامل"
        ws[f"B{row}"].fill = self.style_manager.primary_fill
        ws[f"B{row}"].font = self.style_manager.large_bold_font
        ws[f"B{row}"].alignment = self.style_manager.center_alignment

        # J - Sum of all week totals
        formula = "=" + "+".join([f"J{r}" for r in week_total_rows])
        ws.cell(row=row, column=10).value = formula
        ws.cell(row=row, column=10).fill = self.style_manager.primary_fill
        ws.cell(row=row, column=10).font = self.style_manager.header_font
        ws.cell(row=row, column=10).alignment = self.style_manager.center_alignment

        # Add outline border to grand total
        self.worksheet_helper.outline_range(
            ws,
            start_row=row,
            start_col=2,
            end_row=row,
            end_col=10,
        )

    def write(self, template: SalaryTemplate, use_month_folder: bool = False):
        """
        Write the salary template to an Excel file.

        Args:
            template: SalaryTemplate object containing all data
            use_month_folder: If True, saves to a month folder instead of directly to output
        """
        # Prepare data
        all_employees = self._prepare_employees_list(template)

        # Get attendance filename from path
        import os

        attendance_filename = os.path.basename(template.attendance_file_path)
        borrows_filename = os.path.basename(template.borrows_file_path)

        # Get weeks for the month
        month_weeks = get_weeks(year=template.year, month=template.month)

        # Create workbook and worksheet
        wb = self.create_workbook()
        ws: Worksheet = wb.active
        self.worksheet_helper.configure_worksheet(ws, rtl=True, show_grid=False)

        ws.title = "Salaries"

        # Apply default font to all cells
        self.worksheet_helper.apply_font_to_all(ws, self.style_manager.regular_font)

        # Write sections for each week
        row = 2  # Start from row 2
        week_total_rows = []

        for week_num in range(1, len(month_weeks) + 1):
            row = self._write_week_section(
                ws,
                starting_row=row,
                week_number=week_num,
                employees=all_employees,
                attendance_filename=attendance_filename,
                borrows_filename=borrows_filename,
                template=template,
            )

            # Track the week total row (it's the row before current)
            week_total_rows.append(row - 1)

            # Add 2-row gap between weeks
            row += 2

        # Write grand total
        self._write_grand_total(ws, row, week_total_rows)

        # Set specific column widths for better readability
        ws.column_dimensions["B"].width = 25  # Employee name - wider
        ws.column_dimensions["C"].width = 15  # Daily rate
        ws.column_dimensions["D"].width = 22  # Working days - wider for formula
        ws.column_dimensions["E"].width = 15  # Deductions
        ws.column_dimensions["F"].width = 25  # Working days after deductions - wider
        ws.column_dimensions["G"].width = 18  # Weekly salary
        ws.column_dimensions["H"].width = 25  # Weekly salary with deductions - wider
        ws.column_dimensions["I"].width = 22  # Advances - wider
        ws.column_dimensions["J"].width = 28  # Final salary - wider

        # Set row heights for better readability
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=2)
            # Headers and merged cells need more height
            if cell.alignment and cell.alignment.wrap_text:
                ws.row_dimensions[row_idx].height = 35
            else:
                ws.row_dimensions[row_idx].height = 25

        # Save workbook
        self.save_workbook(
            wb,
            template.year,
            template.month,
            filename_prefix="Salaries_",
            use_month_folder=use_month_folder,
        )
