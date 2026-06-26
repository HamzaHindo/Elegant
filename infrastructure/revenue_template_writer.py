"""
Revenue Template Writer - Generates Excel revenue sheets.
Refactored to use OOP principles with base classes for better reusability.
"""

import calendar
from datetime import date
from typing import List

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from domain.employee import Employee
from domain.revenue_template import RevenueTemplate
from infrastructure.excel_template_writer import ExcelTemplateWriter
from utils.calendar_utils import get_arabic_month, get_weeks, to_arabic_ordinal


class RevenueTemplateWriter(ExcelTemplateWriter):
    """
    Writes revenue templates to Excel files.
    Inherits common functionality from ExcelTemplateWriter.
    """

    def __init__(self, output_directory: str = "./output"):
        """
        Initialize the revenue template writer.

        Args:
            output_directory: Directory where Excel files will be saved
        """
        super().__init__(output_directory)

    def _prepare_employees_list(
        self, employees: list[Employee], stylists
    ) -> list[Employee]:
        """
        Combine employees and stylists into a single list.

        Args:
            employees: List of Employee objects
            stylists: List of Stylist objects

        Returns:
            Combined list of employees and stylists as Employee objects
        """
        stylists_as_employees = [
            Employee(
                name=s.name,
                daily_rate=s.daily_rate,
                working_hours=s.working_hours,
            )
            for s in stylists
        ]
        return employees + stylists_as_employees

    def _write_day_header(self, ws: Worksheet, day: date):
        """
        Write the header section for a day sheet.

        Args:
            ws: Worksheet to write to
            day: Date object for the day
        """
        # Merge cells for title
        ws.merge_cells("B2:P2")
        ws.merge_cells("B3:P3")

        # Title
        ws["B2"] = "الإيرادات اليومية"
        ws["B2"].fill = self.style_manager.primary_fill
        ws["B2"].font = self.style_manager.header_font
        ws["B2"].alignment = ws["B2"].alignment.copy(
            horizontal="center", vertical="center"
        )

        # Date
        ws["B3"] = f"{day.strftime('%A, %d %B %Y')}"
        ws["B3"].fill = self.style_manager.primary_fill
        ws["B3"].font = self.style_manager.header_font
        ws["B3"].alignment = ws["B3"].alignment.copy(
            horizontal="center", vertical="center"
        )

    def _write_income_section(
        self,
        ws: Worksheet,
        stylists,
        payment_methods,
        emp_sheet_name: str,
        pm_sheet_name: str,
    ) -> int:
        """
        Write the income section with invoice rows.

        Args:
            ws: Worksheet to write to
            stylists: List of stylists
            payment_methods: List of payment methods
            emp_sheet_name: Name of employee sheet for dropdown
            pm_sheet_name: Name of payment methods sheet for dropdown

        Returns:
            Number of invoice rows created
        """
        row = 5
        income_headers: List[str] = [
            "عدد الفواتير",
            "المبلغ المحصل",
            "طريقة الدفع",
            "رقم الفاتورة",
            "اسم المصفف",
        ]

        # Write headers
        for i, header in enumerate(income_headers):
            ws[self.cell(row=row, col=i + 2)].value = header
            ws[self.cell(row=row, col=i + 2)].fill = self.style_manager.primary_fill
            ws[self.cell(row=row, col=i + 2)].font = self.style_manager.header_font

        # Empty income rows (30 rows for bills)
        row += 1
        number_of_invoices = 30
        for r in range(row, number_of_invoices + row):
            ws[self.cell(row=r, col=2)].value = r - row + 1  # Bill count
            for c in range(3, 7):
                ws[self.cell(row=r, col=c)].value = ""

        # Payment method dropdown (column D = 4)
        if payment_methods:
            pm_range = f"'{pm_sheet_name}'!$A$3:$A${len(payment_methods)+2}"
            dv_pm = DataValidation(type="list", formula1=pm_range, allow_blank=True)
            ws.add_data_validation(dv_pm)
            dv_pm.add(f"D6:D{number_of_invoices+6}")

        # Stylist dropdown (column F = 6)
        if stylists:
            stylist_range = f"{emp_sheet_name}!$A$1:$A${len(stylists) + 10}"
            dv_stylist = DataValidation(
                type="list", formula1=stylist_range, allow_blank=True
            )
            ws.add_data_validation(dv_stylist)
            dv_stylist.add(f"F6:F{number_of_invoices+6}")

        self.worksheet_helper.outline_range(ws, 5, 2, 36, 6)

        return number_of_invoices

    def _write_totals_section(self, ws: Worksheet, number_of_invoices: int):
        """
        Write the totals section showing collected amounts.

        Args:
            ws: Worksheet to write to
            number_of_invoices: Number of invoice rows
        """
        row = 5
        ws[self.cell(row=row, col=8)].value = "الإجمالى المحصل من العملاء (فيزا او كاش)"
        ws[self.cell(row=row, col=8)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=8)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=8)].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"H{row}:I{row}")

        row += 1
        formula = f"=H{row+2}+I{row+2}"
        ws[self.cell(row=row, col=8)].value = formula
        ws[self.cell(row=row, col=8)].fill = self.style_manager.secondary_fill
        ws[self.cell(row=row, col=8)].font = self.style_manager.large_bold_font
        ws.merge_cells(f"H{row}:I{row}")

        row += 1
        ws[self.cell(row=row, col=8)].value = "فيزا"
        ws[self.cell(row=row, col=8)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=8)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=8)].alignment = Alignment(wrap_text=True)

        ws[self.cell(row=row, col=9)].value = "كاش"
        ws[self.cell(row=row, col=9)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=9)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=9)].alignment = Alignment(wrap_text=True)

        row += 1
        ws[self.cell(row=row, col=8)].value = (
            f'=SUMIF(D6:D{number_of_invoices+6},"فيزا",C6:C{number_of_invoices+6})'
        )
        ws[self.cell(row=row, col=8)].fill = self.style_manager.secondary_fill
        ws[self.cell(row=row, col=9)].value = (
            f'=SUMIF(D6:D{number_of_invoices+6},"كاش",C6:C{number_of_invoices+6})'
        )
        ws[self.cell(row=row, col=9)].fill = self.style_manager.secondary_fill

        row += 1
        ws[self.cell(row=row, col=8)].value = "إجمالي النقدية بالخزينة"
        ws[self.cell(row=row, col=8)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=8)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=8)].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"H{row}:I{row}")

        row += 1
        ws[self.cell(row=row, col=8)].value = ws[self.cell(row=row - 2, col=9)].value
        ws[self.cell(row=row, col=8)].font = self.style_manager.bold_font
        ws[self.cell(row=row, col=8)].alignment = self.style_manager.center_alignment
        ws[self.cell(row=row, col=8)].fill = self.style_manager.secondary_fill
        ws.merge_cells(f"H{row}:I{row}")

        row += 2
        ws[self.cell(row=row, col=8)].value = "الرصيد السابق لللإيراد"
        ws[self.cell(row=row, col=8)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=8)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=8)].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"H{row}:I{row}")

        row += 1
        cell = ws[self.cell(row=row, col=8)]
        if ws.title == "1":
            cell.value = 0
        else:
            prev_sheet = str(int(ws.title) - 1)
            cell.value = f"='{prev_sheet}'!H16"
        cell.font = self.style_manager.bold_font
        cell.alignment = self.style_manager.center_alignment
        cell.fill = self.style_manager.secondary_fill
        ws.merge_cells(f"H{row}:I{row}")

        row += 2
        ws[self.cell(row=row, col=8)].value = "إجمالى الإيراد اليوم بعد المصروفات"
        ws[self.cell(row=row, col=8)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=8)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=8)].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"H{row}:I{row}")

        row += 1
        formula = f"=H6+H13-K{number_of_invoices+1}"
        ws[self.cell(row=row, col=8)].value = formula
        ws[self.cell(row=row, col=8)].font = self.style_manager.large_bold_font
        ws[self.cell(row=row, col=8)].alignment = self.style_manager.center_alignment
        ws[self.cell(row=row, col=8)].fill = self.style_manager.secondary_fill
        ws.merge_cells(f"H{row}:I{row}")

        self.worksheet_helper.outline_range(ws, 5, 8, 10, 9)
        self.worksheet_helper.outline_range(ws, 12, 8, 13, 9)
        self.worksheet_helper.outline_range(ws, 15, 8, 16, 9)

    def _write_expenses_section(
        self,
        ws: Worksheet,
        employees: list[Employee],
        number_of_invoices: int,
        emp_sheet_name: str,
    ):
        """
        Write the expenses section.

        Args:
            ws: Worksheet to write to
            employees: List of employees
            number_of_invoices: Number of invoice rows
            emp_sheet_name: Name of employee sheet for dropdown
        """
        number_of_invoices -= 10
        row = 5
        ws[self.cell(row=row, col=11)].value = "المصروفات من الخزينه"
        ws[self.cell(row=row, col=11)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=11)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=11)].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"K{row}:P{row}")

        row += 1
        expense_headers = ["سلف للموظفين", "مشروبات", "مصاريف نثرية"]
        for i, header in enumerate(expense_headers):
            if header:
                ws[self.cell(row=row, col=11 + 2 * i)].value = header
                ws[self.cell(row=row, col=11 + 2 * i)].fill = (
                    self.style_manager.primary_fill
                )
                ws[self.cell(row=row, col=11 + 2 * i)].font = (
                    self.style_manager.header_font
                )
                ws.merge_cells(
                    f"{get_column_letter(11+2*i)}{row}:{get_column_letter(11+2*i+1)}{row}"
                )

        row += 1
        sub_headers = [
            "اسم الموظف",
            "المبلغ",
            "النوع",
            "القيمة",
            "نوع المصروف",
            "قيمته",
        ]
        for i, header in enumerate(sub_headers):
            ws[self.cell(row=row, col=11 + i)].value = header
            ws[self.cell(row=row, col=11 + i)].fill = self.style_manager.primary_fill
            ws[self.cell(row=row, col=11 + i)].font = self.style_manager.header_font

        row += 1
        # Empty expense rows
        for r in range(row, row + number_of_invoices):
            for c in range(11, 16):
                ws.cell(row=r, column=c).value = ""

        # Employee dropdown in expenses (column L = 12)
        if employees:
            emp_range = f"{emp_sheet_name}!$C$1:$C${len(employees) + 10}"
            dv_emp = DataValidation(type="list", formula1=emp_range, allow_blank=True)
            ws.add_data_validation(dv_emp)
            dv_emp.add(f"K{row}:K{row+number_of_invoices-1}")

        row += number_of_invoices
        expense_total_headers = [
            "إجمالي سلف الموظفين",
            "إجمالي المشروبات",
            "إجمالي المصاريف النثرية",
        ]
        for i, header in enumerate(expense_total_headers):
            if header:
                ws[self.cell(row=row, col=11 + 2 * i)].value = header
                ws[self.cell(row=row, col=11 + 2 * i)].fill = (
                    self.style_manager.primary_fill
                )
                ws[self.cell(row=row, col=11 + 2 * i)].font = (
                    self.style_manager.header_font
                )
                ws.merge_cells(
                    f"{get_column_letter(11+2*i)}{row}:{get_column_letter(11+2*i+1)}{row}"
                )

        # Expense totals
        row += 1
        for i, header in enumerate(expense_total_headers):
            if header:
                col = 11 + 2 * i
                ws[self.cell(row=row, col=col)].value = (
                    f"=SUM({get_column_letter(col+1)}8:{get_column_letter(col+1)}{8-1+number_of_invoices})"
                )
                ws[self.cell(row=row, col=11 + 2 * i)].font = (
                    self.style_manager.bold_font
                )
                ws.merge_cells(
                    f"{get_column_letter(11+2*i)}{row}:{get_column_letter(11+2*i+1)}{row}"
                )

        row += 1
        ws[self.cell(row=row, col=11)].value = (
            "إجمالى جميع المصروفات من الخزينه (سلف الموظفين+المشروبات+المصاريف النثرية)"
        )
        ws[self.cell(row=row, col=11)].alignment = Alignment(wrap_text=True)
        ws[self.cell(row=row, col=11)].font = self.style_manager.header_font
        ws[self.cell(row=row, col=11)].fill = self.style_manager.primary_fill
        ws.merge_cells(f"K{row}:P{row}")

        row += 1
        ws[self.cell(row=row, col=11)].value = f"=K{row-2}+M{row-2}+O{row-2}"
        ws[self.cell(row=row, col=11)].font = self.style_manager.bold_font
        ws.merge_cells(f"K{row}:P{row}")

        self.worksheet_helper.outline_range(ws, 5, 11, 11 + number_of_invoices, 16)

    def _write_targets_section(self, ws: Worksheet, stylists, number_of_invoices: int):
        """
        Write the targets section for stylists.

        Args:
            ws: Worksheet to write to
            stylists: List of stylists
            number_of_invoices: Number of invoice rows (adjusted)
        """
        row = 40
        ws.merge_cells(f"G{row}:J{row}")
        ws[self.cell(row=row, col=7)].value = "تارجت اليوم للمصففين"
        ws[self.cell(row=row, col=7)].fill = self.style_manager.primary_fill
        ws[self.cell(row=row, col=7)].font = self.style_manager.header_font

        row += 1
        target_headers = [
            "إسم الموظف",
            "التارجت اليومى",
            "التارجت المحقق",
            "النسبة",
        ]
        for i, header in enumerate(target_headers):
            ws[self.cell(row=row, col=7 + i)].value = header
            ws[self.cell(row=row, col=7 + i)].fill = self.style_manager.primary_fill
            ws[self.cell(row=row, col=7 + i)].font = self.style_manager.header_font

        # Stylist rows
        row += 1
        for i, stylist in enumerate(stylists):
            ws[self.cell(row=row + i, col=7)].value = stylist.name
            ws[self.cell(row=row + i, col=8)].value = stylist.daily_target
            ws[self.cell(row=row + i, col=8)].font = (
                self.style_manager.target_value_font
            )
            ws[self.cell(row=row + i, col=9)].value = (
                f'=SUMIF(F6:F31,"{stylist.name}",C6:C{6+number_of_invoices})'
            )
            ws[self.cell(row=row + i, col=9)].font = (
                self.style_manager.target_value_font
            )
            ws[self.cell(row=row + i, col=10)].value = (
                f"={self.cell(row=row+i, col=9)} / 10"
            )
            ws[self.cell(row=row + i, col=10)].font = (
                self.style_manager.target_percentage_font
            )

        self.worksheet_helper.outline_range(ws, 40, 7, 40 + len(stylists) + 1, 10)

    def _write_day_sheet(
        self,
        wb,
        day: date,
        stylists,
        employees,
        payment_methods,
        sheet_index: int,
        emp_sheet_name: str,
        pm_sheet_name: str,
    ):
        """
        Write a complete day sheet.

        Args:
            wb: Workbook object
            day: Date object for the day
            stylists: List of stylists
            employees: List of employees
            payment_methods: List of payment methods
            sheet_index: Index of the sheet (0 for first sheet)
            emp_sheet_name: Name of employee sheet
            pm_sheet_name: Name of payment methods sheet
        """
        # Create or use worksheet
        if sheet_index == 0:
            ws: Worksheet = wb.active
        else:
            ws = wb.create_sheet()

        # Prepare employees list
        employees = self._prepare_employees_list(employees, stylists)

        # Configure worksheet
        ws.title = f"{day.day}"
        self.worksheet_helper.configure_worksheet(ws, rtl=True, show_grid=False)

        # Apply default font
        self.worksheet_helper.apply_font_to_all(ws, self.style_manager.bold_font)

        # Write sections
        self._write_day_header(ws, day)
        number_of_invoices = self._write_income_section(
            ws, stylists, payment_methods, emp_sheet_name, pm_sheet_name
        )
        self._write_totals_section(ws, number_of_invoices)
        self._write_expenses_section(ws, employees, number_of_invoices, emp_sheet_name)
        self._write_targets_section(ws, stylists, number_of_invoices - 10)

        # Apply final formatting
        self.worksheet_helper.center_range(
            ws,
            start_row=1,
            start_col=1,
            end_row=ws.max_row,
            end_col=ws.max_column,
        )
        self.worksheet_helper.auto_fit_sheet(ws)

    def _write_employees(self, wb, employees, stylists):
        """
        Write the employees lookup sheet.

        Args:
            wb: Workbook object
            employees: List of employees
            stylists: List of stylists
        """
        ws = wb.create_sheet()
        ws.title = "الموظفين"
        self.worksheet_helper.configure_worksheet(ws, rtl=True, show_grid=False)

        # Column A -> Stylist names
        # Column B -> Daily targets
        row = 1
        for stylist in stylists:
            ws.cell(row=row, column=1).value = stylist.name
            ws.cell(row=row, column=2).value = stylist.daily_target
            row += 1

        # Column C -> Employees + Stylists
        row = 1
        all_people = employees + stylists

        for person in all_people:
            ws.cell(row=row, column=3).value = person.name
            row += 1

    def _write_payment_methods(self, wb, payment_methods):
        """
        Write the payment methods lookup sheet.

        Args:
            wb: Workbook object
            payment_methods: List of payment methods
        """
        ws = wb.create_sheet()
        ws.title = "طرق الدفع"
        self.worksheet_helper.configure_worksheet(ws, rtl=True, show_grid=False)

        # Title
        ws.merge_cells("A1:B1")
        ws["A1"] = "طرق الدفع"
        ws["A1"].fill = self.style_manager.primary_fill
        ws["A1"].font = self.style_manager.header_font
        ws["A1"].alignment = ws["A1"].alignment.copy(
            horizontal="center", vertical="center"
        )

        # Headers
        row = 2
        headers = ["طريقة الدفع", "الوصف"]
        for i, header in enumerate(headers):
            ws.cell(row=row, column=i + 1).value = header
            ws.cell(row=row, column=i + 1).fill = self.style_manager.primary_fill
            ws.cell(row=row, column=i + 1).font = self.style_manager.header_font

        # Data
        descriptions = {
            "فيزا": "دفع إلكتروني",
            "كاش": "دفع نقدي",
        }
        row = 3
        for pm in payment_methods:
            ws.cell(row=row, column=1).value = pm
            ws.cell(row=row, column=2).value = descriptions.get(pm, "")
            row += 1

        # Apply borders and font
        self.worksheet_helper.add_borders_to_range(
            ws, max_row_idx=row - 1, max_col_idx=2
        )
        for row_cells in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row_cells:
                if not cell.font or cell.font.name != "Dubai":
                    cell.font = self.style_manager.regular_font

        self.worksheet_helper.auto_fit_sheet(ws)

    def _write_total_borrows_sheet(self, wb, template: RevenueTemplate):
        """
        Write the total borrows sheet.
        Clones the attendance table structure with one row per day,
        week totals, and a month total.
        """
        ws = wb.create_sheet()
        ws.title = "كشف السلف"
        self.worksheet_helper.configure_worksheet(ws, rtl=True, show_grid=False)
        self.worksheet_helper.apply_font_to_all(ws, self.style_manager.arial_regular_font)

        employees = self._prepare_employees_list(template.employees, template.stylists)
        num_employees = len(employees)
        last_col = 2 + num_employees

        # Header: A2:B3 merged = "التاريخ"
        ws.merge_cells("A2:B3")
        ws["A2"] = "التاريخ"
        ws["A2"].fill = self.style_manager.primary_fill
        ws["A2"].font = self.style_manager.arial_bold_font
        ws["A2"].alignment = self.style_manager.center_alignment

        # Title: C2 to last_col merged
        ws.merge_cells(f"C2:{get_column_letter(last_col)}2")
        ws["C2"] = (
            f"كشف السلف للموظفين عن شهر {get_arabic_month(template.month)} {template.year}"
        )
        ws["C2"].fill = self.style_manager.primary_fill
        ws["C2"].font = self.style_manager.arial_bold_font
        ws["C2"].alignment = self.style_manager.center_alignment

        # Employee names in row 3
        for i, emp in enumerate(employees):
            col = 3 + i
            ws.cell(row=3, column=col).value = emp.name
            ws.cell(row=3, column=col).font = self.style_manager.dubai_bold_font
            ws.cell(row=3, column=col).alignment = self.style_manager.center_alignment

        # Days and week totals
        row = 4
        month_weeks = get_weeks(year=template.year, month=template.month)
        each_week_total_row = []

        for j, week in enumerate(month_weeks):
            week_start_row = row
            for day in week:
                d = date(year=template.year, month=template.month, day=day)
                ws.merge_cells(f"A{row}:B{row}")
                ws.cell(row=row, column=1).value = d
                ws.cell(row=row, column=1).font = self.style_manager.dubai_bold_font
                ws.cell(row=row, column=1).number_format = "dddd, mmmm d, yyyy"
                ws.cell(row=row, column=1).alignment = self.style_manager.center_alignment
                # Columns C onwards are left empty for manual entry
                for i in range(num_employees):
                    col = 3 + i
                    ws.cell(row=row, column=col).alignment = self.style_manager.center_alignment
                row += 1

            # Week total row
            ws.merge_cells(f"A{row}:B{row}")
            ws.cell(row=row, column=1).value = f"نهاية الأسبوع {to_arabic_ordinal(j + 1)}"
            ws.cell(row=row, column=1).fill = self.style_manager.primary_fill
            ws.cell(row=row, column=1).font = self.style_manager.arial_bold_font
            ws.cell(row=row, column=1).alignment = self.style_manager.center_alignment

            for i in range(num_employees):
                col = 3 + i
                start_cell = self.cell(col=col, row=week_start_row)
                end_cell = self.cell(col=col, row=row - 1)
                ws.cell(row=row, column=col).value = f"=SUM({start_cell}:{end_cell})"
                ws.cell(row=row, column=col).fill = self.style_manager.primary_fill
                ws.cell(row=row, column=col).font = self.style_manager.arial_bold_font
                ws.cell(row=row, column=col).alignment = self.style_manager.center_alignment

            each_week_total_row.append(row)
            row += 1

        # Month total row
        ws.merge_cells(f"A{row}:B{row}")
        ws.cell(row=row, column=1).value = "نهاية الشهر"
        ws.cell(row=row, column=1).fill = self.style_manager.primary_fill
        ws.cell(row=row, column=1).font = self.style_manager.arial_bold_font
        ws.cell(row=row, column=1).alignment = self.style_manager.center_alignment

        for i in range(num_employees):
            col = 3 + i
            formula = "="
            for m, total_row in enumerate(each_week_total_row):
                formula += self.cell(col=col, row=total_row)
                if m != len(each_week_total_row) - 1:
                    formula += "+"
            ws.cell(row=row, column=col).value = formula
            ws.cell(row=row, column=col).fill = self.style_manager.primary_fill
            ws.cell(row=row, column=col).font = self.style_manager.arial_bold_font
            ws.cell(row=row, column=col).alignment = self.style_manager.center_alignment

        # Grand total for all employees (2 big rows)
        row += 1
        ws.merge_cells(f"A{row}:{get_column_letter(last_col)}{row}")
        ws.cell(row=row, column=1).value = "إجمالي سلف جميع الموظفين"
        ws.cell(row=row, column=1).fill = self.style_manager.primary_fill
        ws.cell(row=row, column=1).font = self.style_manager.arial_bold_font
        ws.cell(row=row, column=1).alignment = self.style_manager.center_alignment

        month_total_row = row - 1
        row += 1
        ws.merge_cells(f"A{row}:{get_column_letter(last_col)}{row}")
        first_emp_col = get_column_letter(3)
        last_emp_col = get_column_letter(last_col)
        ws.cell(row=row, column=1).value = (
            f"=SUM({first_emp_col}{month_total_row}:{last_emp_col}{month_total_row})"
        )
        ws.cell(row=row, column=1).fill = self.style_manager.secondary_fill
        ws.cell(row=row, column=1).font = self.style_manager.large_bold_font
        ws.cell(row=row, column=1).alignment = self.style_manager.center_alignment

        # Apply borders and final formatting
        days_in_month = calendar.monthrange(template.year, template.month)[1]
        max_used_col = last_col
        max_used_row = row
        self.worksheet_helper.add_borders_to_range(
            ws, max_col_idx=max_used_col, max_row_idx=max_used_row
        )
        self.worksheet_helper.auto_fit_sheet(ws)
        ws.freeze_panes = "A4"

    def write(self, template: RevenueTemplate, use_month_folder: bool = False):
        """
        Write the revenue template to an Excel file.

        Args:
            template: RevenueTemplate object containing all data
            use_month_folder: If True, saves to a month folder instead of directly to output
        """
        wb = self.create_workbook()

        emp_sheet_name = "الموظفين"
        pm_sheet_name = "طرق الدفع"

        # Write day sheets
        for i, day in enumerate(template.days):
            self._write_day_sheet(
                wb,
                day,
                template.stylists,
                template.employees,
                template.payment_methods,
                i,
                emp_sheet_name,
                pm_sheet_name,
            )

        # Save workbook with custom prefix

        # Write lookup sheets first so they exist when day sheets reference them
        self._write_employees(wb, template.employees, template.stylists)
        self._write_payment_methods(wb, template.payment_methods)
        self._write_total_borrows_sheet(wb, template)
        self.save_workbook(
            wb,
            template.year,
            template.month,
            filename_prefix="Revenue_",
            use_month_folder=use_month_folder,
        )
