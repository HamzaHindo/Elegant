"""
Excel Style Manager - Centralized styling for Excel workbooks.
Provides reusable style definitions and application methods.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelStyleManager:
    """Manages all Excel styling operations including fonts, fills, borders, and alignment."""

    # Color constants
    # Primary Colors
    COLOR_NAVY = "1F3A5F"  # Main headers
    COLOR_NAVY_DARK = "152A45"  # Section headers
    COLOR_SKY = "DCE6F1"  # Light background
    COLOR_WHITE = "FFFFFF"

    # Accent Colors
    COLOR_GREEN = "2E7D32"  # Success
    COLOR_RED = "C62828"  # Warnings
    COLOR_GOLD = "B8860B"  # Highlights
    COLOR_BLUE_DARK = "00008B"  # Dark blue for targets
    COLOR_GREEN_DARK = "023020"  # Dark green for percentages

    # Text Colors
    COLOR_TEXT = "1F2937"  # Dark gray
    COLOR_TEXT_LIGHT = "FFFFFF"

    # Border Colors
    COLOR_BLACK = "000000"  # Borders

    def __init__(self):
        """Initialize the style manager with predefined styles."""
        self._init_fills()
        self._init_fonts()
        self._init_borders()
        self._init_alignments()

    def _init_fills(self):
        """Initialize fill patterns."""
        self.primary_fill = PatternFill(
            start_color=self.COLOR_NAVY,
            end_color=self.COLOR_NAVY,
            fill_type="solid",
        )

        self.secondary_fill = PatternFill(
            start_color=self.COLOR_SKY,
            end_color=self.COLOR_SKY,
            fill_type="solid",
        )

    def _init_fonts(self):
        """Initialize font styles."""
        self.header_font = Font(
            name="Dubai",
            size=12,
            bold=True,
            color=self.COLOR_WHITE,
        )

        self.regular_font = Font(
            name="Dubai",
            size=11,
            color=self.COLOR_TEXT,
        )

        self.bold_font = Font(
            name="Dubai",
            size=11,
            bold=True,
            color=self.COLOR_TEXT,
        )

        self.large_bold_font = Font(
            name="Dubai",
            size=22,
            bold=True,
            color=self.COLOR_NAVY,
        )

        self.arial_regular_font = Font(
            name="Arial",
            size=11,
            color=self.COLOR_TEXT,
        )

        # Changed from arial to dubai
        self.arial_bold_font = Font(
            name="Dubai",
            size=14,
            bold=True,
            color=self.COLOR_WHITE,
        )

        self.dubai_bold_font = Font(
            name="Dubai",
            size=14,
            bold=True,
            color=self.COLOR_NAVY,
        )

        # Target section fonts
        self.target_value_font = Font(
            name="Dubai",
            size=11,
            bold=True,
            color=self.COLOR_BLUE_DARK,
        )

        self.target_percentage_font = Font(
            name="Dubai",
            size=11,
            bold=True,
            color=self.COLOR_GREEN_DARK,
        )

    def _init_borders(self):
        """Initialize border styles."""
        self.thin_side = Side(border_style="thin", color=self.COLOR_BLACK)
        self.medium_side = Side(border_style="medium", color=self.COLOR_BLACK)
        self.no_border_side = Side(border_style=None)

        self.visible_border = Border(
            left=self.thin_side,
            right=self.thin_side,
            top=self.thin_side,
            bottom=self.thin_side,
        )
        self.no_border = Border(
            left=self.no_border_side,
            right=self.no_border_side,
            top=self.no_border_side,
            bottom=self.no_border_side,
        )

    def _init_alignments(self):
        """Initialize alignment styles."""
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.wrap_alignment = Alignment(wrap_text=True)
        self.center_wrap_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def get_custom_font(self, name="Dubai", size=11, bold=False, color=None) -> Font:
        """Create a custom font with specified parameters."""
        if color is None:
            color = self.COLOR_TEXT
        return Font(name=name, size=size, bold=bold, color=color)

    def get_custom_border(self, style="thin", color=None, sides=None) -> Border:
        """
        Create a custom border with specified parameters.

        Args:
            style: Border style (thin, medium, thick, etc.)
            color: Border color
            sides: List of sides to apply border ['left', 'right', 'top', 'bottom']
                   If None, applies to all sides
        """
        if color is None:
            color = self.COLOR_BLACK
        side = Side(border_style=style, color=color)
        no_side = Side(border_style=None)

        if sides is None:
            sides = ["left", "right", "top", "bottom"]

        return Border(
            left=side if "left" in sides else no_side,
            right=side if "right" in sides else no_side,
            top=side if "top" in sides else no_side,
            bottom=side if "bottom" in sides else no_side,
        )
