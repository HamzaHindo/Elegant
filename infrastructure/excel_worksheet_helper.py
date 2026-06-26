"""
Excel Worksheet Helper - Common worksheet operations.
Provides utility methods for cell references, auto-fitting, borders, and more.
"""

from copy import copy

from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from infrastructure.excel_style_manager import ExcelStyleManager


class ExcelWorksheetHelper:
    """Helper class for common Excel worksheet operations."""

    def __init__(self, style_manager: ExcelStyleManager):
        """
        Initialize the worksheet helper.

        Args:
            style_manager: ExcelStyleManager instance for styling operations
        """
        self.style_manager = style_manager

    @staticmethod
    def cell(col: int, row: int) -> str:
        """
        Convert column and row numbers to Excel cell reference.

        Args:
            col: Column number (1-indexed)
            row: Row number (1-indexed)

        Returns:
            Excel cell reference (e.g., "A1", "B2")
        """
        return f"{get_column_letter(col)}{row}"

    def configure_worksheet(
        self, ws: Worksheet, rtl: bool = True, show_grid: bool = False
    ):
        """
        Configure basic worksheet settings.

        Args:
            ws: Worksheet to configure
            rtl: Right-to-left layout (default True for Arabic)
            show_grid: Show gridlines (default False)
        """
        ws.sheet_view.rightToLeft = rtl
        ws.sheet_view.showGridLines = show_grid

    def auto_fit_columns(
        self, ws: Worksheet, skip_formulas: bool = True, skip_top_rows: int = 2
    ):
        """
        Auto-fit column widths based on content.

        Args:
            ws: Worksheet to adjust
            skip_formulas: Skip cells with formulas when calculating width
            skip_top_rows: Number of top rows to skip (usually headers)
        """
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.value is None:
                    continue

                value = str(cell.value)

                # Skip formulas if requested
                if skip_formulas and value.startswith("="):
                    continue

                # Skip top rows if requested
                if row_idx <= skip_top_rows:
                    continue

                max_length = max(max_length, len(value))

            # Calculate adjusted width with bounds
            adjusted_width = min(max(max_length, 8), 40) + 10
            ws.column_dimensions[column_letter].width = adjusted_width

    def auto_fit_rows(self, ws: Worksheet, max_height: int = 60):
        """
        Auto-fit row heights based on content.

        Args:
            ws: Worksheet to adjust
            max_height: Maximum row height to prevent excessive heights
        """
        for row_idx in range(1, ws.max_row + 1):
            max_lines = 1

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.value:
                    lines = str(cell.value).count("\n") + 1
                    max_lines = max(max_lines, lines)

            # Calculate height with bounds
            ws.row_dimensions[row_idx].height = min(max_lines * 25, max_height) + 10

    def auto_fit_sheet(
        self,
        ws: Worksheet,
        skip_formulas: bool = True,
        skip_top_rows: int = 2,
        max_row_height: int = 60,
    ):
        """
        Auto-fit both columns and rows.

        Args:
            ws: Worksheet to adjust
            skip_formulas: Skip cells with formulas when calculating column width
            skip_top_rows: Number of top rows to skip for column width calculation
            max_row_height: Maximum row height
        """
        self.auto_fit_columns(ws, skip_formulas, skip_top_rows)
        self.auto_fit_rows(ws, max_row_height)

    def remove_all_borders(self, ws: Worksheet):
        """
        Remove borders from all cells in the worksheet.

        Args:
            ws: Worksheet to modify
        """
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = self.style_manager.no_border

    def add_borders_to_range(self, ws: Worksheet, max_row_idx: int, max_col_idx: int):
        """
        Add borders to a specific range of cells.

        Args:
            ws: Worksheet to modify
            max_row_idx: Maximum row index for borders
            max_col_idx: Maximum column index for borders
        """
        # Hide Excel's background gridlines
        ws.sheet_view.showGridLines = False

        # Apply borders to specified range
        for row_cells in ws.iter_rows(
            min_row=1, max_row=max_row_idx, min_col=1, max_col=max_col_idx
        ):
            for cell in row_cells:
                cell.border = self.style_manager.visible_border

    def outline_range(
        self,
        ws: Worksheet,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
    ):
        """
        Add outline borders to a range (thicker borders on edges).

        Args:
            ws: Worksheet to modify
            start_row: Starting row
            start_col: Starting column
            end_row: Ending row
            end_col: Ending column
        """
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)

                cell.border = Border(
                    left=medium if col == start_col else thin,
                    right=medium if col == end_col else thin,
                    top=medium if row == start_row else thin,
                    bottom=medium if row == end_row else thin,
                )

    def center_range(
        self,
        ws: Worksheet,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
    ):
        """
        Center align all cells in a range.

        Args:
            ws: Worksheet to modify
            start_row: Starting row
            start_col: Starting column
            end_row: Ending row
            end_col: Ending column
        """
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)

                alignment = copy(cell.alignment)
                alignment.horizontal = "center"
                alignment.vertical = "center"
                alignment.wrap_text = True

                cell.alignment = alignment

    def apply_font_to_all(self, ws: Worksheet, font):
        """
        Apply a font to all cells in the worksheet.

        Args:
            ws: Worksheet to modify
            font: Font object to apply
        """
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.font = font
